"""Build a privacy-minimised daily alert from the latest redacted staff-folder XLSX.

Only initials/staff references, checklist labels and date status are emitted.
"""
from datetime import date, datetime
from pathlib import Path
import sys
from openpyxl import load_workbook

WORKBOOK = Path(__file__).with_name("latest-staff-folder.xlsx")
DATE_COLUMNS = {"Car MOT": "MOT", "Car insurance": "Car insurance"}


def is_red(cell):
    """Return True only for an explicit solid Excel red fill (#FF0000)."""
    fill = cell.fill
    return (
        fill.fill_type == "solid"
        and fill.fgColor.type == "rgb"
        and fill.fgColor.rgb in {"FFFF0000", "FF0000"}
    )


def display_date(value):
    return value.strftime("%-d %b %Y")


def main():
    if not WORKBOOK.exists():
        print("Staff-folder daily alert could not run: no latest spreadsheet is loaded.")
        return

    try:
        book = load_workbook(WORKBOOK, data_only=True, read_only=True)
        sheet = book.active
    except Exception as exc:
        print(f"Staff-folder daily alert could not read the latest spreadsheet: {exc}")
        return

    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    try:
        staff_col = next(i + 1 for i, header in enumerate(headers) if str(header).strip().upper() in {"STAFF", "EMPLOYEE"})
    except StopIteration:
        print("Staff-folder daily alert could not find a STAFF/EMPLOYEE column.")
        return

    today = date.today()
    red_by_staff = {}
    expired_by_staff = {}
    invalid_dates = []

    for row in range(2, sheet.max_row + 1):
        staff = sheet.cell(row, staff_col).value
        if staff in (None, "", "Last audited", "By who"):
            continue
        staff = str(staff).strip()
        red_items = []
        expired_items = []
        for col, header in enumerate(headers, start=1):
            if not header or col == staff_col:
                continue
            cell = sheet.cell(row, col)
            label = str(header).strip()
            if is_red(cell):
                red_items.append(label)
            if label in DATE_COLUMNS:
                if isinstance(cell.value, datetime):
                    expiry = cell.value.date()
                    if expiry < today:
                        expired_items.append(f"{DATE_COLUMNS[label]} expired {display_date(expiry)}")
                elif cell.value not in (None, "", "NA", "N/A"):
                    invalid_dates.append(f"{staff}: invalid {DATE_COLUMNS[label]} date ({cell.value})")
        if red_items:
            red_by_staff[staff] = red_items
        if expired_items:
            expired_by_staff[staff] = expired_items

    lines = [f"Staff-folder daily alert — {today.strftime('%-d %B %Y')}"]
    if not red_by_staff and not expired_by_staff and not invalid_dates:
        lines.append("No red cells, expired MOTs/insurance dates, or invalid vehicle dates in the latest spreadsheet.")
    else:
        if red_by_staff:
            total_red = sum(len(items) for items in red_by_staff.values())
            lines.append(f"RED CELLS: {total_red} across {len(red_by_staff)} staff record(s).")
            for staff, items in red_by_staff.items():
                preview = ", ".join(items[:4])
                more = f" (+{len(items) - 4} more)" if len(items) > 4 else ""
                lines.append(f"• {staff}: {preview}{more}")
        if expired_by_staff:
            lines.append("EXPIRED DATES:")
            for staff, items in expired_by_staff.items():
                lines.append(f"• {staff}: {'; '.join(items)}")
        if invalid_dates:
            lines.append("DATA TO FIX:")
            lines.extend(f"• {item}" for item in invalid_dates)
    lines.append("Green is acceptable. Date alerts are included only after the expiry date has passed.")
    print("\n".join(lines))


if __name__ == "__main__":
    main()
